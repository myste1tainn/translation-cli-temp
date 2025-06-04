from openpyxl import load_workbook
from tabulate import tabulate


def excel_to_markdown(path, sheet_name=None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Find the max range
    max_row = ws.max_row
    max_col = ws.max_column

    # Build merged cell lookup
    merged_lookup = {}
    for merged in ws.merged_cells.ranges:
        min_col = merged.min_col
        min_row = merged.min_row
        value = ws.cell(row=min_row, column=min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                merged_lookup[(row, col)] = value

    # Fill in values with merged support
    table = []
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = merged_lookup.get((row, col), cell.value)
            row_data.append("" if value is None else str(value))
        table.append(row_data)

    # Trim trailing empty rows/columns
    while table and all(cell == "" for cell in table[-1]):
        table.pop()

    # Convert to markdown
    markdown = tabulate(table, tablefmt="github")
    with open("output.md", "w") as file:
        file.write(markdown)
