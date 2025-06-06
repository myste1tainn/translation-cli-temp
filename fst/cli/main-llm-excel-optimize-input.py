import pandas as pd
import openpyxl
from openpyxl.cell.cell import (
    TYPE_STRING,
    TYPE_NUMERIC,
    TYPE_BOOL,
    TYPE_FORMULA,
    TYPE_NULL,
    TYPE_ERROR,
)
from googletrans import Translator
import asyncio
import json
import re
import uuid
from aiclient import ai_chat

translator = Translator()

usage_token = 0

to_translate_words = []
to_translate_cells = []

jargon_list_str = """
{
        "สินทรัพย์": "Assets",
        "สินทรัพย์หมุนเวียน": "Current assets",
        "เงินสดและรายการเทียบเท่าเงินสด": "Cash and cash equivalents",
        "เงินลงทุนชั่วคราว": "Current investments",
        "ลูกหนี้การค้าและลูกหนี้อื่น": "Trade and other receivables",
        "เงินให้กู้ยืมระยะสั้น": "Current loans and advances",
        "สินค้าคงเหลือ": "Inventories",
        "สินทรัพย์หมุนเวียนอื่น": "Other current assets",
        "รวมสินทรัพย์หมุนเวียน": "Total current assets",
        "สินทรัพย์ไม่หมุนเวียน": "Non-current assets",
        "เงินลงทุนเผื่อขาย": "Investments held as available-for-sale",
        "เงินลงทุนในบริษัทร่วม": "Investments in associates",
        "เงินลงทุนในบริษัทย่อย": "Investments in subsidiaries",
        "เงินลงทุนในการร่วมค้า": "Investments in joint ventures",
        "เงินลงทุนระยะยาวอื่น": "Other non-current investments",
        "เงินให้กู้ยืมระยะยาว": "Non-current loans and advances",
        "อสังหาริมทรัพย์เพื่อการลงทุน": "Investment property",
        "สินทรัพย์ไม่หมุนเวียนที่ถือไว้เพื่อขาย": "Non-current assets classified as held for sale",
        "ที่ดิน อาคารและอุปกรณ์": "Property, plant and equipment",
        "สินทรัพย์ไม่มีตัวตน": "Intangible assets",
        "สินทรัพย์ไม่หมุนเวียนอื่น": "Other non-current assets",
        "รวมสินทรัพย์ไม่หมุนเวียน": "Total non-current assets",
        "รวมสินทรัพย์": "Total assets",
        "หนี้สินและส่วนของผู้ถือหุ้น": "Liabilities and equity",
        "หนี้สินหมุนเวียน": "Current liabilities",
        "เงินเบิกเกินบัญชีและเงินกู้ยืมระยะสั้นจากสถาบันการเงิน": "Bank overdrafts and short-term borrowing from financial institutions",
        "เจ้าหนี้การค้าและเจ้าหนี้อื่น": "Trade and other payables",
        "ส่วนของหนี้สินระยะยาวที่ถึงกำหนดชำระภายในหนึ่งปี": "Current portion of long-term liabilities",
        "เงินกู้ยืมระยะสั้น": "Short-term borrowings",
        "ภาษีเงินได้ค้างจ่าย": "Current income tax payable",
        "ประมาณการหนี้สินระยะสั้น": "Short-term provisions",
        "หนี้สินหมุนเวียนอื่น": "Other current liabilities",
        "รวมหนี้สินหมุนเวียน": "Total current liabilities",
        "หนี้สินไม่หมุนเวียน": "Non-current liabilities",
        "เงินกู้ยืมระยะยาว": "Long-term borrowings",
        "ประมาณการหนี้สินผลประโยชน์พนักงาน": "Employee benefit obligation",
        "ประมาณการหนี้สินระยะยาว": "Long-term provisions",
        "หนี้สินไม่หมุนเวียนอื่น": "Other non-current liabilities",
        "รวมหนี้สินไม่หมุนเวียน": "Total non-current liabilities",
        "รวมหนี้สิน": "Liabilities",
        "ส่วนของผู้ถือหุ้น": "Shareholders' equity",
        "ทุนเรือนหุ้น": "Share capital",
        "ทุนจดทะเบียน": "Authorized share capital",
        "หุ้นบุริมสิทธิ": "Preference shares",
        "จำนวนหุ้น": "Number of shares",
        "มูลค่าต่อหุ้น": "Par value",
        "หุ้นสามัญ": "Ordinary shares",
        "ทุนที่ชำระแล้ว": "Paid-up share capital",
        "ส่วนเกินมูลค่าหุ้น": "Share premium account",
        "ส่วนเกินมูลค่าหุ้นบุริมสิทธิ": "Share premium account-preference shares",
        "ส่วนเกินมูลค่าหุ้นสามัญ": "Share premium account-ordinary shares",
        "กำไร (ขาดทุน) สะสม": "Retained earnings",
        "จัดสรรแล้ว": "Appropriated",
        "ทุนสำรองตามกฎหมาย": "Legal reserves",
        "อื่นๆ": "Others",
        "ยังไม่ได้จัดสรร": "Unappropriated",
        "องค์ประกอบอื่นของส่วนของผู้ถือหุ้น": "Other components of shareholders' equity",
        "รวมส่วนของผู้ถือหุ้น": "Total shareholders' equity",
        "รวมหนี้สินและส่วนของผู้ถือหุ้น": "Total liabilities and shareholders' equity",
        "งบกำไรขาดทุน": "Statement of income",
        "รายได้": "Revenue",
        "รายได้จากการขายหรือการให้บริการ": "Revenues from sales or revenues from services",
        "รายได้ดอกเบี้ย": "Interest income",
        "รายได้ค่าก่อสร้าง": "Contract revenue",
        "รายได้อื่น": "Other incomes",
        "รวมรายได้": "Total revenue",
        "ค่าใช้จ่าย": "Expenses",
        "ต้นทุนขายหรือต้นทุนการให้บริการ": "Costs of sale of goods and rendering of services",
        "ต้นทุนการก่อสร้าง": "Contract costs",
        "ค่าใช้จ่ายในการขาย": "Selling expense",
        "ค่าใช้จ่ายในการบริหาร": "Administrative expense",
        "ค่าใช้จ่ายอื่น": "Other expenses",
        "รวมค่าใช้จ่าย": "Total expenses",
        "กำไร (ขาดทุน) ก่อนต้นทุนทางการเงินและค่าใช้จ่ายภาษีเงินได้": "Profit (loss) before finance costs and income tax expense",
        "ต้นทุนทางการเงิน": "Finance costs",
        "กำไร (ขาดทุน) ก่อนค่าใช้จ่ายภาษีเงินได้": "Profit (loss) before income tax expense",
        "ค่าใช้จ่ายภาษีเงินได้": "Income tax expense",
        "กำไร (ขาดทุน) สุทธิ": "Net profit (loss)",
        "งบแสดงการเปลี่ยนแปลงส่วนของผู้ถือหุ้น": "Statement of changes in shareholders' equity",
        "ยอดคงเหลือต้นงวด": "Equity balance at beginning of period",
        "ผลกระทบของการเปลี่ยนแปลงนโยบายการบัญชี": "Financial effects of changes in accounting policies",
        "ผลสะสมจากการแก้ไขข้อผิดพลาดทางการบัญชี": "Financial effects of corrections of accounting errors",
        "ยอดคงเหลือที่ปรับปรุงแล้ว": "Balance-restated",
        "การเปลี่ยนแปลงในส่วนของผู้ถือหุ้น": "Changes in shareholders' equity",
        "การเพิ่ม (ลด) หุ้นสามัญ": "Increase (decrease)-ordinary shares",
        "การเพิ่ม (ลด) หุ้นบุริมสิทธิ": "Increase (decrease)-preference shares",
        "เงินปันผล": "Dividends",
        "ผลต่างของอัตราแลกเปลี่ยนจากการแปลงค่างบการเงิน": "Exchange difference on translating financial statements",
        "ผลกำไร (ขาดทุน) จากการวัดมูลค่าเงินลงทุนเผื่อขาย": "Gains (losses) on remeasuring available-for-sale investments",
        "ผลกำไร (ขาดทุน) ที่ยังไม่เกิดขึ้นจริงอื่น": "Other unrealized gains (losses)",
        "รวมการเปลี่ยนแปลงส่วนของผู้ถือหุ้น": "Total changes in shareholders' equity",
        "ยอดคงเหลือปลายงวด": "Equity balance at ending of period"
}"""

# Generate a unique request ID
request_id = str(uuid.uuid4())


async def google_translate_text(value, dest_lang="en"):
    if isinstance(value, str) and value.strip():
        res = await translator.translate(value, dest=dest_lang)
        return res.text
    return value


def gen_ai_translate_text(cell, value, company, is_merged):
    if (
        isinstance(value, str)
        and bool(re.fullmatch(r"[\d,.\-\s]+", value)) == False
        and value.strip()
    ):

        if is_merged:
            return value
        else:
            global to_translate_cells
            global to_translate_words

            to_translate_cells.append((cell.row, cell.column))
            to_translate_words.append(value)
            return ""

    return value


def open_ai_translate(value, company):
    response = ai_chat(
        messages=[
            {
                "role": "system",
                "content": "You are a financial service translator assistant.",
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": f"""
Your job is to translate input in Thai to output in English. Please be aware that the content you're translating is from financial statement of a company called: {company}.

You should also use the provided jargon word list to ensure accurate translation. If the input matches any word or phrase from the jargon list, use the predefined translation. If no exact match is found, translate the input based on financial terminology.

jargon word list: {jargon_list_str}

input value: {value}

If the term is not found in the jargon word list, translate based on financial context while maintaining accuracy.

Please give the answer in json format. Respond only with valid JSON. Do not write an introduction or summary:
{{
    "translated_value": "value1"
}}
Example: 
For the request where the the input value: "ยอดคงเหลือต้นงวด"
The response should be:
{{
 "translated_value": "Equity balance at beginning of period"
}}
                    """,
                    }
                ],
            },
        ]
    )

    global usage_token
    usage_token += response.usage.completion_tokens

    json_response = response.choices[0].message.content

    if json_response is None:
        return open_ai_translate(value, company)

    response = json.loads(json_response)
    result = response.get("translated_value")
    print(f"translating {value}, translated {result}")
    return result


def open_ai_translate_words(sheet, value, company):
    response = ai_chat(
        [
            {
                "role": "system",
                "content": "You are a financial service translator assistant.",
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": f"""
Your job is to translate input in Thai to output in English. Please be aware that the content you're translating is from financial statement of a company called: {company}.

You should also use the provided jargon word list to ensure accurate translation. If the input matches any word or phrase from the jargon list, use the predefined translation. If no exact match is found, translate the input based on financial terminology.

jargon word list: {jargon_list_str}

input value: {value}

If the term is not found in the jargon word list, translate based on financial context while maintaining accuracy.

Please give the answer in json format. Respond only with valid JSON. Do not write an introduction or summary:
{{
    "translated_value": ["value1","value2"]
}}
Example: 
For the request where the the input value: ["ยอดคงเหลือต้นงวด"]
The response should be:
{{
 "translated_value": ["Equity balance at beginning of period"]
}}
                    """,
                    }
                ],
            },
        ]
    )

    global usage_token
    usage_token += response.usage.completion_tokens

    json_response = response.choices[0].message.content

    if json_response is None:
        return open_ai_translate_words(sheet, value, company)

    print(f"json_response: {json_response}")
    response = json.loads(json_response)
    result = response.get("translated_value")
    print(f"translating {value}, translated {result}")

    global to_translate_cells
    global to_translate_words

    for index, translated_value in enumerate(result):
        translated_cell = to_translate_cells[index]
        cell = sheet.cell(row=translated_cell[0], column=translated_cell[1])
        cell.value = translated_value

    to_translate_cells = []
    to_translate_words = []


def convert_be_to_ad_in_text(text):
    """Find all B.E. years in the text and convert them to A.D."""
    # Find all B.E. years using regex
    be_years = re.findall(r"\bพ\.ศ\.\s*(\d{4})\b", text, flags=re.IGNORECASE)

    # Replace each B.E. year with its A.D. equivalent
    for be_year in be_years:
        ad_year = int(be_year) - 543
        newText = text.replace(f"พ.ศ. {be_year}", f"Year {ad_year}")
        if newText != text:
            text = newText
        else:
            text = text.replace(f"พ.ศ.{be_year}", f"Year {ad_year}")

    return text


def get_translated_text(cell, company, jargon_list_json, is_merged):
    jargon_word = jargon_list_json.get(cell.value, None)
    if jargon_word != None:
        translated_text = jargon_word
    else:
        translated_text = gen_ai_translate_text(cell, cell.value, company, is_merged)

    return translated_text


async def convert_using_openpyxl(file_path, output_path, company):
    startTime = pd.Timestamp.now()
    print(f"##### Start time: {startTime}")

    jargon_list_json = json.loads(jargon_list_str)

    wb = openpyxl.load_workbook(file_path)
    sheetCount = 0
    for sheet in wb.worksheets:
        sheetCount = sheetCount + 1

        print(f"================= Sheet {sheetCount}: {sheet} | {sheet.sheet_state} ")

        if sheet.sheet_state == sheet.SHEETSTATE_HIDDEN:
            continue

        merged_ranges = sheet.merged_cells.ranges
        i = 0
        for row in sheet.iter_rows():
            print(f"Row no: {i}")
            i = i + 1
            for cell in row:
                # Skip if the cell contains a formula
                if (
                    cell.data_type == TYPE_FORMULA
                    or cell.data_type == TYPE_NULL
                    or cell.data_type == TYPE_ERROR
                    or cell.data_type == TYPE_BOOL
                    or cell.data_type == TYPE_NUMERIC
                ):
                    continue
                if cell.value is None or cell.value == "":
                    continue
                elif cell.value == "บาท":
                    cell.value = "Baht"
                    continue
                elif cell.data_type == TYPE_STRING:
                    if re.fullmatch(r"[0-9.,]+", cell.value):
                        continue
                    elif cell.value.startswith("พ.ศ."):
                        cell.value = convert_be_to_ad_in_text(cell.value)
                        continue

                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        top_left_cell = sheet[merged_range.start_cell.coordinate]
                        if cell.coordinate == top_left_cell.coordinate and isinstance(
                            cell.value, str
                        ):
                            cell.value = get_translated_text(
                                cell, company, jargon_list_json, True
                            )
                        break
                else:
                    cell.value = get_translated_text(
                        cell, company, jargon_list_json, False
                    )

        open_ai_translate_words(sheet, to_translate_words, company)

    wb.save(output_path)

    endTime = pd.Timestamp.now()
    print(f"##### End time: {endTime}")
    print(f"######### Total time: {endTime - startTime} | Total token: {usage_token}")


if __name__ == "__main__":
    print("Start converting\n\n")
    # Load Excel
    file_path = "./data/gpsc/FINANCIAL_STATEMENTS.XLSX"
    output_path = "translated_gpsc_sheet3-gpt-4o-batch.xlsx"  # claude-3-5-sonnet-v2 #gemini-pro-1.5 #gpt-4o
    asyncio.run(
        convert_using_openpyxl(
            file_path, output_path, "Global Power Synergy Public Company Limited"
        )
    )
    print("Done converting the file\n\n")
