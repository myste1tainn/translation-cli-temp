import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def process_financial_statements():
    # Load all sheets from the translated Excel file
    financial_data_dict = pd.read_excel(
        "PTTEP_FINANCIAL_STATEMENTS_EN_Cut version.xlsx", sheet_name=None
    )

    # Load the original workbook to retain styles
    original_wb = load_workbook("PTTEP_FINANCIAL_STATEMENTS_EN_Cut version.xlsx")

    # Create a new workbook to save the modified data
    modified_wb = load_workbook("PTTEP_FINANCIAL_STATEMENTS_EN_Cut version.xlsx")

    for sheet_name, financial_data in financial_data_dict.items():
        # Extract all values from column 0
        column_0_values = financial_data.iloc[:, 0].tolist()

        # Create a new list with "modified_" prepended to each value
        modified_values = ["modified_" + str(value) for value in column_0_values]

        # Convert the list to a pandas Series to ensure compatibility
        modified_values_series = pd.Series(modified_values, index=financial_data.index)

        # Insert the new Series as a new column immediately after column 0
        financial_data.insert(1, "Modified Column 0", modified_values_series)

        # Get the original and modified worksheets
        original_ws = original_wb[sheet_name]
        modified_ws = modified_wb[sheet_name]

        # Clear the existing data in the modified worksheet
        for row in modified_ws.iter_rows(
            min_row=2, max_row=modified_ws.max_row, max_col=modified_ws.max_column
        ):
            for cell in row:
                # Only clear the cell if it is not part of a merged cell
                if not any(
                    cell.coordinate in merged_range
                    for merged_range in modified_ws.merged_cells.ranges
                ):
                    cell.value = None

        # Append the modified data to the new worksheet
        for r_idx, row in enumerate(
            dataframe_to_rows(financial_data, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                # Only set the cell value if it is not part of a merged cell
                if not any(
                    modified_ws.cell(row=r_idx, column=c_idx).coordinate in merged_range
                    for merged_range in modified_ws.merged_cells.ranges
                ):
                    modified_ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the new workbook with the original formatting
    modified_wb.save("Modified_Financial_Statements.xlsx")
