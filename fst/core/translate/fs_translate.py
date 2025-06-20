import os
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import openpyxl
from openpyxl.cell.cell import (
    # This is probably wrong because using this causes a warning in the LSP everywhere
    # Cell,
    TYPE_STRING,
    TYPE_NUMERIC,
    TYPE_BOOL,
    TYPE_FORMULA,
    TYPE_NULL,
    TYPE_ERROR,
)
from googletrans import Translator
import json
import re
import uuid
from typing import Any
from fst.ai.client import ai_chat
from fst.core.translate.jargon_list import jargon_list_str
from markitdown import MarkItDown

translator = Translator()

usage_token = 0

to_translate_words: list[str] = []
to_translate_cells: list[tuple[int, int]] = []


# Generate a unique request ID
request_id = str(uuid.uuid4())

fs_translate_debug = False


def gen_ai_translate_text(
    cell: Any,
    value: str,
    _: str,  # This is the company name for context in translation, commented out becaus it's not in used right now
    is_merged: bool,
) -> str:
    if (
        isinstance(value, str)
        and bool(re.fullmatch(r"[\d,.\-\s]+", value)) == False
        and value.strip()
    ):
        global to_translate_cells
        global to_translate_words
        if is_merged:
            # return value

            to_translate_cells.append((cell.row, cell.column))
            to_translate_words.append(value)
            return ""
        else:
            # global to_translate_cells
            # global to_translate_words

            to_translate_cells.append((cell.row, cell.column))
            to_translate_words.append(value)
            return ""

    return value


def open_ai_translate_words(sheet: Worksheet, value: list[str], company: str) -> None:
    response = ai_chat(
        [
            {
                "role": "system",
                "content": "You are a financial service translator assistant.",
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": f"""
Your job is to translate input in Thai to output in English. Please be aware that the content you're translating is from financial statement of a company called: {company}.

You should also use the provided jargon word list to ensure accurate translation. If the input matches any word or phrase from the jargon list, use the predefined translation. If no exact match is found, translate the input based on financial terminology.

jargon word list: {jargon_list_str}

input value: {value}

If the term is not found in the jargon word list, translate based on financial context while maintaining accuracy.

Please give the answer in json format. Respond only with valid JSON. Do not write an introduction or summary:
{{
    "translated_value": ["value1","value2"]
}}
Example: 
For the request where the the input value: ["ยอดคงเหลือต้นงวด"]
The response should be:
{{
 "translated_value": ["Equity balance at beginning of period"]
}}
                    """,
                    }
                ],
            },
        ]
    )

    global usage_token
    if response.usage is not None:
        usage_token += response.usage.completion_tokens

    json_response = response.choices[0].message.content

    if json_response is None:
        return open_ai_translate_words(sheet, value, company)

    _print(f"json_response: {json_response}")
    json_object: dict[str, Any] = json.loads(json_response)
    # Mark as Any for now, because don't know what type is expected
    # on the line that says `cell.value = translated_value`
    result: list[Any] = json_object.get("translated_value") or []
    _print(f"translating {value}, translated {result}")

    global to_translate_cells
    global to_translate_words

    for index, translated_value in enumerate(result):
        translated_cell = to_translate_cells[index]
        cell = sheet.cell(row=translated_cell[0], column=translated_cell[1])
        cell.value = translated_value

    to_translate_cells = []
    to_translate_words = []


def convert_be_to_ad_in_text(text: str) -> str:
    """Find all B.E. years in the text and convert them to A.D."""
    # Find all B.E. years using regex
    be_years: list[str] = re.findall(r"\bพ\.ศ\.\s*(\d{4})\b", text, flags=re.IGNORECASE)

    # Replace each B.E. year with its A.D. equivalent
    for be_year in be_years:
        ad_year = int(be_year) - 543
        newText = text.replace(f"พ.ศ. {be_year}", f"Year {ad_year}")
        if newText != text:
            text = newText
        else:
            text = text.replace(f"พ.ศ.{be_year}", f"Year {ad_year}")

    return text


def get_translated_text(
    cell: Any,
    company: str,
    jargon_list_json: dict[str, str],
    is_merged: bool,
) -> str:
    jargon_word: str | None = jargon_list_json.get(cell.value, None)
    if jargon_word != None:
        _print(f"##### matched jargon word")
        translated_text = jargon_word
    else:
        _print(f"##### matched jargon word, try with AI")
        translated_text = gen_ai_translate_text(cell, cell.value, company, is_merged)

    _print(f"##### translated_text {jargon_word}")
    return translated_text


async def translate(file_path: str, output_path: str, company: str) -> tuple[str, str]:
    startTime = pd.Timestamp.now()
    _print(f"##### Start time: {startTime}")

    jargon_list_json: dict[str, str] = json.loads(jargon_list_str)

    wb = openpyxl.load_workbook(file_path)
    sheetCount = 0
    for sheet in wb.worksheets:
        sheetCount = sheetCount + 1

        _print(f"================= Sheet {sheetCount}: {sheet} | {sheet.sheet_state} ")

        if sheet.sheet_state == sheet.SHEETSTATE_HIDDEN:
            continue

        merged_ranges = sheet.merged_cells.ranges
        i = 0
        for row in sheet.iter_rows():
            _print(f"Row no: {i}")
            i = i + 1
            for cell in row:
                # Skip if the cell contains a formula
                if (
                    cell.data_type == TYPE_FORMULA
                    or cell.data_type == TYPE_NULL
                    or cell.data_type == TYPE_ERROR
                    or cell.data_type == TYPE_BOOL
                    or cell.data_type == TYPE_NUMERIC
                ):
                    continue

                if cell.value is None or cell.value == "":
                    continue
                elif cell.value == "บาท":
                    cell.value = "Baht"
                    continue
                elif cell.data_type == TYPE_STRING and isinstance(cell.value, str):
                    v = str(cell.value)
                    if re.fullmatch(r"[0-9.,]+", v):
                        continue
                    elif v.startswith("พ.ศ."):
                        cell.value = convert_be_to_ad_in_text(v)
                        continue
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        _print(
                            f"################## cells found in merged_ranges value is {cell.value}"
                        )

                        # NOTE: The code below that accesses `coordinate` works just fine, but the LSP doesn't seem to recognize it.
                        # force down cast to shut up the LSP warning
                        # from typing import cast

                        top_left_cell: Any = sheet[merged_range.start_cell.coordinate]
                        _print(f"################## top_left_cell {cell.value}")
                        _print(
                            f"##########,####### {cell.coordinate},{top_left_cell.coordinate}"
                        )
                        if cell.coordinate == top_left_cell.coordinate and isinstance(
                            cell.value, str
                        ):
                            cell.value = get_translated_text(
                                cell, company, jargon_list_json, True
                            )
                            _print(f"################## translated text = {cell.value}")
                        break
                else:
                    _print("normal")
                    cell.value = get_translated_text(
                        cell, company, jargon_list_json, False
                    )

        open_ai_translate_words(sheet, to_translate_words, company)

    wb.save(output_path)
    md = MarkItDown().convert(output_path)
    md_path = f"out/{os.path.basename(output_path).replace('.xlsx', '.md')}"
    with open(md_path, "w") as file:
        file.write(md.text_content)

    endTime = pd.Timestamp.now()
    _print(f"##### End time: {endTime}")
    _print(f"######### Total time: {endTime - startTime} | Total token: {usage_token}")
    return md_path, output_path


def _print(text: str) -> None:
    if not fs_translate_debug:
        return
    print(text)
